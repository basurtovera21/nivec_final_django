from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from datetime import date
import openpyxl

from core.models import (Universidad, Campus, Carrera, UsuarioDeSistema, PerfilDocente, PerfilEstudiante, PerfilAdministrativo, MallaCurricular, UnidadCurricular, PeriodoDeNivelacion, Paralelo, Horario, CohorteDeMatricula, MatriculaParalelo, ConsolidadoAcademico, EvaluacionAcademica, InformeGeneral)
from core.forms import (DocumentoMTN, FormularioUniversidad, FormularioCampus, FormularioCarrera, FormularioUsuarioDeSistema, FormularioPerfilDocente, FormularioPerfilEstudiante, FormularioPerfilAdministrativo, FormularioMallaCurricular, FormularioUnidadCurricular, FormularioPeriodoDeNivelacion, FormularioParalelo, FormularioHorario, FormularioCohorteDeMatricula, FormularioEvaluacionAcademica)
from clases.enums.estado_de_matricula import EstadoDeMatricula
from clases.enums.estado_de_aprobacion import EstadoDeAprobacion
from clases.enums.estado_de_periodo import EstadoDePeriodo
from clases.enums.estado_de_informe import EstadoDeInforme
from clases.enums.formato_de_exportacion import FormatoDeExportacion

#Autenticación
def iniciar_sesion(request):
    #Correo institucional y contraseña.
    if request.method == "POST":
        correo_institucional = request.POST.get("correo_institucional")
        contrasena = request.POST.get("contrasena")
        usuario_de_sistema = authenticate(request, username = correo_institucional, password = contrasena)
        if usuario_de_sistema is not None:
            login(request, usuario_de_sistema)
            return redirect("panel_principal")
        else:
            messages.error(request, "Correo institucional o contraseña no válidos.")
    return render(request, "autenticacion/iniciar_sesion.html")


@login_required
def cerrar_sesion(request):
    logout(request)
    return redirect("iniciar_sesion")


@login_required
def panel_principal(request):
    #Redirección al panel correspondiente según el perfil del usuario.
    usuario_de_sistema = request.user
    if hasattr(usuario_de_sistema, "perfil_administrativo"): #hasattr; verifica si posee el atributo especificado.
        perfil = usuario_de_sistema.perfil_administrativo.perfil_administrativo #Tipo de perfil
        if perfil in ("Director de dirección de admisión y nivelación",
                      "Coordinador de dirección de admisión y nivelación"):
            return redirect("panel_dan")
        
        elif perfil == "Coordinador de unidad académica":
            return redirect("panel_coordinador_ua")
        
        else:
            return redirect("panel_administrativo")
        
    elif hasattr(usuario_de_sistema, "perfil_docente"):
        return redirect("panel_docente")
    
    elif hasattr(usuario_de_sistema, "perfil_estudiante"):
        return redirect("panel_estudiante")
    
    return render(request, "autenticacion/sin_perfil.html")




#Panel por rol
@login_required
def panel_dan(request):
    periodos_de_nivelacion = PeriodoDeNivelacion.objects.all().order_by("-anio")
    consolidados_academicos = ConsolidadoAcademico.objects.all()
    return render(request, "dan/panel_dan.html", {
        "periodos_de_nivelacion": periodos_de_nivelacion,
        "consolidados_academicos": consolidados_academicos,
    })


@login_required
def panel_coordinador_ua(request):
    paralelos = Paralelo.objects.all().order_by("periodo_de_nivelacion")
    docentes = PerfilDocente.objects.filter(estado_de_vinculacion = "Activo")
    return render(request, "coordinador_ua/panel_coordinador_ua.html", {
        "paralelos": paralelos,
        "docentes": docentes,
    })


@login_required
def panel_docente(request):
    try:
        perfil_docente = request.user.perfil_docente
        paralelos = Paralelo.objects.filter(docente_responsable = perfil_docente)
        return render(request, "docente/panel_docente.html", {
            "perfil_docente": perfil_docente,
            "paralelos": paralelos,
        })
    except PerfilDocente.DoesNotExist:
        return redirect("panel_principal")


@login_required
def panel_estudiante(request):
    try:
        perfil_estudiante = request.user.perfil_estudiante
        matriculas = MatriculaParalelo.objects.filter(estudiante = perfil_estudiante).select_related("paralelo", "paralelo__unidad_curricular")
        evaluaciones_academicas = EvaluacionAcademica.objects.filter(estudiante = perfil_estudiante)
        return render(request, "estudiante/panel_estudiante.html", {
            "perfil_estudiante": perfil_estudiante,
            "matriculas": matriculas,
            "evaluaciones_academicas": evaluaciones_academicas,
        })
    except PerfilEstudiante.DoesNotExist:
        return redirect("panel_principal")


@login_required
def panel_administrativo(request):
    universidades = Universidad.objects.all()
    return render(request, "administrativo/panel_administrativo.html", {
        "universidades": universidades,
    })
    
    
    
    
#Entidades
@login_required
def listar_universidades(request):
    universidades = Universidad.objects.all()
    return render(request, "entidades/listar_universidades.html", {
        "universidades": universidades
    })


@login_required
def registrar_universidad(request):
    if request.method == "POST": #Verificar si el formulario fue enviado.
        formulario_universidad = FormularioUniversidad(request.POST, request.FILES)
        if formulario_universidad.is_valid():
            formulario_universidad.save()
            messages.success(request, "La universidad ha sido registrada correctamente.")
            return redirect("listar_universidades")
    else:
        formulario_universidad = FormularioUniversidad() #Crea un formulario vacío
    return render(request, "entidades/formulario_universidad.html", {
        "formulario_universidad": formulario_universidad,
        "titulo": "Registrar universidad",
    })


@login_required
def listar_campus(request):
    campus = Campus.objects.all().select_related("universidad")
    return render(request, "entidades/listar_campus.html", {"campus": campus})


@login_required
def registrar_campus(request):
    if request.method == "POST":
        formulario_campus = FormularioCampus(request.POST)
        if formulario_campus.is_valid():
            formulario_campus.save()
            messages.success(request, "El campus ha sido registrado correctamente.")
            return redirect("listar_campus")
    else:
        formulario_campus = FormularioCampus()
    return render(request, "entidades/formulario_campus.html", {
        "formulario_campus": formulario_campus,
        "titulo": "Registrar campus",
    })


@login_required
def listar_carreras(request):
    carreras = Carrera.objects.all().select_related("campus")
    return render(request, "entidades/listar_carreras.html", {"carreras": carreras})


@login_required
def registrar_carrera(request):
    if request.method == "POST":
        formulario_carrera = FormularioCarrera(request.POST)
        if formulario_carrera.is_valid():
            formulario_carrera.save()
            messages.success(request, "La carrera ha sido registrada correctamente.")
            return redirect("listar_carreras")
    else:
        formulario_carrera = FormularioCarrera()
    return render(request, "entidades/formulario_carrera.html", {
        "formulario_carrera": formulario_carrera,
        "titulo": "Registrar carrera",
    })




#Administración de usuarios
@login_required
def listar_docentes(request):
    docentes = PerfilDocente.objects.all().select_related("usuario_de_sistema")
    return render(request, "usuarios/listar_docentes.html", {"docentes": docentes})


@login_required
def registrar_docente(request):
    #Crear UsuarioDeSistema y PerfilDocente.    
    if request.method == "POST":
        formulario_usuario_de_sistema = FormularioUsuarioDeSistema(request.POST)
        formulario_perfil_docente  = FormularioPerfilDocente(request.POST)
        if formulario_usuario_de_sistema.is_valid() and formulario_perfil_docente.is_valid():
            usuario_de_sistema = formulario_usuario_de_sistema.save(commit = False) #Crear objeto pero no se guarda
            usuario_de_sistema.set_password(formulario_usuario_de_sistema.cleaned_data["contrasena"]) #Cifrar contraseña
            usuario_de_sistema.save() #Guardar en la base de datos
            docente = formulario_perfil_docente.save(commit = False)
            docente.usuario_de_sistema = usuario_de_sistema #Relacionar Docente y UsuarioDeSistema
            docente.save()
            messages.success(request, "El docente ha sido registrado correctamente.")
            return redirect("listar_docentes")
        
    else:
        formulario_usuario_de_sistema = FormularioUsuarioDeSistema()
        formulario_perfil_docente = FormularioPerfilDocente()
        
    return render(request, "usuarios/formulario_docente.html", {
        "formulario_usuario_de_sistema": formulario_usuario_de_sistema,
        "formulario_perfil_docente" : formulario_perfil_docente,
        "titulo": "Registrar docente",
    })


@login_required
def listar_estudiantes(request):
    estudiantes = PerfilEstudiante.objects.all().select_related("usuario_de_sistema", "carrera_registrada", "campus_registrado")
    return render(request, "usuarios/listar_estudiantes.html", {
        "estudiantes": estudiantes
    })


@login_required
def registrar_estudiante(request):
    if request.method == "POST":
        formulario_usuario_de_sistema = FormularioUsuarioDeSistema(request.POST)
        formulario_perfil_estudiante = FormularioPerfilEstudiante(request.POST)
        if formulario_usuario_de_sistema.is_valid() and formulario_perfil_estudiante.is_valid():
            usuario_de_sistema = formulario_usuario_de_sistema.save(commit = False)
            usuario_de_sistema.set_password(formulario_usuario_de_sistema.cleaned_data["contrasena"])
            usuario_de_sistema.save()
            estudiante = formulario_perfil_estudiante.save(commit = False)
            estudiante.usuario_de_sistema = usuario_de_sistema
            estudiante.save()
            messages.success(request, "El estudiante ha sido registrado correctamente.")
            return redirect("listar_estudiantes")
        
    else:
        formulario_usuario_de_sistema = FormularioUsuarioDeSistema()
        formulario_perfil_estudiante  = FormularioPerfilEstudiante()
        
    return render(request, "usuarios/formulario_estudiante.html", {
        "formulario_usuario_de_sistema": formulario_usuario_de_sistema,
        "formulario_perfil_estudiante": formulario_perfil_estudiante,
        "titulo": "Registrar estudiante",
    })




#Estructura académica
@login_required
def listar_periodos(request):
    periodos_de_nivelacion = PeriodoDeNivelacion.objects.all().order_by("-anio")
    return render(request, "academico/listar_periodos.html", {"periodos": periodos_de_nivelacion})


@login_required
def registrar_periodo_de_nivelacion(request):
    if request.method == "POST":
        formulario_periodo_de_nivelacion = FormularioPeriodoDeNivelacion(request.POST)
        if formulario_periodo_de_nivelacion.is_valid():
            formulario_periodo_de_nivelacion.save()
            messages.success(request, "El periodo de nivelación ha sido registrado correctamente.")
            return redirect("listar_periodos")
        
    else:
        formulario_periodo_de_nivelacion = FormularioPeriodoDeNivelacion()
        
    return render(request, "academico/formulario_periodo.html", {
        "formulario_periodo_de_nivelacion": formulario_periodo_de_nivelacion,
        "titulo": "Registrar periodo de nivelación",
    })


@login_required
def listar_mallas_curriculares(request):
    mallas_curriculares = MallaCurricular.objects.all().select_related("carrera")
    return render(request, "academico/listar_mallas.html", {"mallas": mallas_curriculares})


@login_required
def registrar_malla_curricular(request):
    if request.method == "POST":
        formulario_malla_curricular = FormularioMallaCurricular(request.POST)
        if formulario_malla_curricular.is_valid():
            formulario_malla_curricular.save()
            messages.success(request, "La malla curricular ha sido registrada correctamente.")
            return redirect("listar_mallas")
        
    else:
        formulario_malla_curricular = FormularioMallaCurricular()
        
    return render(request, "academico/formulario_malla.html", {
        "formulario": formulario_malla_curricular,
        "titulo": "Registrar malla curricular",
    })


@login_required
def listar_unidades_curriculares(request):
    unidades_curriculares = UnidadCurricular.objects.all().select_related("malla_curricular")
    return render(request, "academico/listar_unidades.html", {"unidades": unidades_curriculares})


@login_required
def registrar_unidades_curriculares(request):
    if request.method == "POST":
        formulario_unidad_curricular = FormularioUnidadCurricular(request.POST)
        if formulario_unidad_curricular.is_valid():
            formulario_unidad_curricular.save()
            messages.success(request, "La unidad curricular ha sido registrada correctamente.")
            return redirect("listar_unidades")
        
    else:
        formulario_unidad_curricular = FormularioUnidadCurricular()
        
    return render(request, "academico/formulario_unidad.html", {
        "formulario_unidad_curricular": formulario_unidad_curricular,
        "titulo": "Registrar unidad curricular",
    })


@login_required
def listar_paralelos(request):
    paralelos = Paralelo.objects.all().select_related("periodo_de_nivelacion", "unidad_curricular", "docente_responsable")
    return render(request, "academico/listar_paralelos.html", {"paralelos": paralelos})


@login_required
def registrar_paralelo(request):
    if request.method == "POST":
        formulario_paralelo = FormularioParalelo(request.POST)
        if formulario_paralelo.is_valid():
            formulario_paralelo.save()
            messages.success(request, "El paralelo ha sido registrado correctamente.")
            return redirect("listar_paralelos")
    else:
        formulario_paralelo = FormularioParalelo()
        
    return render(request, "academico/formulario_paralelo.html", {
        "formulario_paralelo": formulario_paralelo,
        "titulo": "Registrar paralelo",
    })


@login_required
def crear_horario(request):
    if request.method == "POST":
        formulario_horario = FormularioHorario(request.POST)
        if formulario_horario.is_valid():
            formulario_horario.save()
            messages.success(request, "El horario ha sido registrado correctamente.")
            return redirect("listar_paralelos")
        
    else:
        formulario_horario = FormularioHorario()
        
    return render(request, "academico/formulario_horario.html", {
        "formulario_horario": formulario_horario,
        "titulo": "Registrar horario",
    })




#Procesar MTN y Consolidado académico
@login_required
def procesar_mtn(request):
    #Crear UsuarioDeSistema y PerfilEstudiante por cada fila válida.
    #Registros totales en ConsolidadoAcademico.
    #identificacion, nombres, apellidos, correo_institucional, carrera, campus, jornada, registro_de_cupo
    if request.method == "POST":
        formulario_documento_mtn = DocumentoMTN(request.POST, request.FILES)
        if formulario_documento_mtn.is_valid():
            documento_mtn = request.FILES["documento_mtn"]
            documento_mtn_en_memoria = openpyxl.load_workbook(documento_mtn)
            hoja_activa = documento_mtn_en_memoria.active
            registros_totales = 0
            registros_validos = 0
            registros_observados = 0
            filas_observadas = []

            for fila in hoja_activa.iter_rows(min_row = 2, values_only = True):
                registros_totales += 1
                try:
                    identificacion = fila[0]
                    nombres = fila[1]
                    apellidos = fila[2]
                    correo_institucional = fila[3]
                    carrera_nombre = fila[4]
                    campus_nombre = fila[5]
                    jornada = fila[6]
                    registro_de_cupo = fila[7]
                    
                    if not all([identificacion, nombres, apellidos, correo_institucional, carrera_nombre, campus_nombre, jornada, registro_de_cupo]):
                        raise ValueError("No se ha completado toda la información requerida.")

                    #Buscar Carrera y Campus registrados.
                    carrera = Carrera.objects.filter(nombre__iexact = str(carrera_nombre)).first()
                    campus  = Campus.objects.filter(nombre__iexact = str(campus_nombre)).first()

                    if not carrera or not campus:
                        raise ValueError(f"Carrera o Campus no registrado ({carrera_nombre}/{campus_nombre}).")

                    #Evitar duplicado
                    if UsuarioDeSistema.objects.filter(identificacion = str(identificacion)).exists():
                        raise ValueError(f"El número de identificación ya ha sido registrado ({identificacion}).")

                    #Crear UsuarioDeSistema
                    usuario_de_sistema = UsuarioDeSistema.objects.create_user(
                        correo_institucional = str(correo_institucional),
                        password = str(identificacion),  # Contraseña inicial = identificación
                        identificacion = str(identificacion),
                        nombres = str(nombres),
                        apellidos = str(apellidos),
                    )

                    #Crear PerfilEstudiante
                    PerfilEstudiante.objects.create(
                        usuario_de_sistema = usuario_de_sistema,
                        identificador_institucional = str(identificacion),
                        numero_de_matricula = str(identificacion),
                        jornada = str(jornada),
                        registro_de_cupo = str(registro_de_cupo),
                        carrera_registrada = carrera,
                        campus_registrado = campus,
                        estado_de_matricula = EstadoDeMatricula.ASPIRANTE.value,
                    )
                    registros_validos += 1

                except Exception as error:
                    registros_observados += 1
                    filas_observadas.append(f"Fila {registros_totales + 1}: {error}")

            #Registrar ConsolidadoAcademimco
            periodo_de_nivelacion = PeriodoDeNivelacion.objects.filter(estado = EstadoDePeriodo.PLANIFICACION.value).first() #Buscar primer periodo en planificación
            if periodo_de_nivelacion:
                #Actualizar o crear registro (evitar duplicado)
                ConsolidadoAcademico.objects.update_or_create(
                    #Si existe el consolidado con el periodo seleccionado
                    periodo_academico = periodo_de_nivelacion,
                    #Valores creados o actualizados
                    defaults = {
                        "fecha_de_corte": date.today(),
                        "total_cupos_aceptados": registros_validos,
                        "registros_totales": registros_totales,
                        "registros_validos": registros_validos,
                        "registros_observados": registros_observados,
                    }
                )

            messages.success(
                request,
                f"Documento procesado, número de registros; {registros_validos} válidos, {registros_observados} observados de {registros_totales} registros."
            )
            if filas_observadas:
                for fila in filas_observadas:
                    messages.warning(request, fila) #Errores

            return redirect("listar_estudiantes")
    else:
        formulario_documento_mtn = DocumentoMTN()

    return render(request, "dan/cargar_mtn.html", {
        "formulario": formulario_documento_mtn,
        "titulo": "Procesar Matriz de Tercer Nivel (MTN)",
    })




#Evaluaciones e informe
@login_required
def listar_evaluaciones(request):
    evaluaciones_academicas = EvaluacionAcademica.objects.all().select_related("estudiante", "unidad_curricular")
    return render(request, "academico/listar_evaluaciones.html", {
        "evaluaciones_academicas": evaluaciones_academicas
    })


@login_required
def registrar_evaluacion(request):
    if request.method == "POST":
        formulario_evaluacion_academica = FormularioEvaluacionAcademica(request.POST)
        if formulario_evaluacion_academica.is_valid():
            evaluacion = formulario_evaluacion_academica.save(commit = False)
            #Calcular nota final
            parcial_1 = evaluacion.calificacion_parcial_1
            parcial_2 = evaluacion.calificacion_parcial_2
            evaluacion.nota_final = round((parcial_1 + parcial_2)/2, 2)
            # Verificar aprobación
            unidad_curricular = evaluacion.unidad_curricular
            if evaluacion.porcentaje_asistencia < unidad_curricular.porcentaje_minimo_asistencia:
                evaluacion.estado_de_aprobacion = EstadoDeAprobacion.REPROBADO.value
                evaluacion.observacion = "Reprobado por porcentaje de asistencia insuficiente."
                
            elif evaluacion.nota_final >= unidad_curricular.criterio_de_aprobacion:
                evaluacion.estado_de_aprobacion = EstadoDeAprobacion.APROBADO.value
                
            else:
                evaluacion.estado_de_aprobacion = EstadoDeAprobacion.REPROBADO.value
                evaluacion.observacion = "Reprobado por calificación insuficiente."
                
            evaluacion.save()
            messages.success(request, "La evaluación académica ha sido registrada correctamente.")
            return redirect("listar_evaluaciones")
    else:
        formulario_evaluacion_academica = FormularioEvaluacionAcademica()
    return render(request, "academico/formulario_evaluacion.html", {
        "formulario": formulario_evaluacion_academica,
        "titulo": "Registrar evaluación académica",
    })


@login_required
def crear_informe_general(request, periodo_id):
    periodo = get_object_or_404(PeriodoDeNivelacion, pk = periodo_id) #Si no existe devuelve el error 404
    evaluaciones_academicas = EvaluacionAcademica.objects.filter(
        unidad_curricular__malla_curricular__carrera__campus__universidad = periodo.universidad #Evaluaciones pertenecientes de la universidad del periodo seleccionado.
    ).select_related("estudiante", "unidad_curricular")

    formato = request.GET.get("formato", "excel")
    if formato == "excel":
        return _exportar_informe_excel(periodo, evaluaciones_academicas)
    else:
        return _exportar_informe_pdf(periodo, evaluaciones_academicas)


def _exportar_informe_excel(periodo, evaluaciones):
    documento_en_memoria = openpyxl.Workbook()
    hoja_activa  = documento_en_memoria.active
    hoja_activa.title = f"Informe general {periodo.periodo}"

    # Encabezados
    hoja_activa.append([
        "Número de identificación", "Nombres", "Apellidos",
        "Carrera", "Campus", "Jornada",
        "Unidad curricular", "Parcial 1", "Parcial 2",
        "Nota final", "Porcentaje de asistencia (%)", "Estado"
    ])

    for evaluacion in evaluaciones:
        estudiante_actual = evaluacion.estudiante
        hoja_activa.append([
            estudiante_actual.usuario_de_sistema.identificacion,
            estudiante_actual.usuario_de_sistema.nombres,
            estudiante_actual.usuario_de_sistema.apellidos,
            estudiante_actual.carrera_registrada.nombre,
            estudiante_actual.campus_registrado.nombre,
            estudiante_actual.jornada,
            evaluacion.unidad_curricular.nombre,
            evaluacion.calificacion_parcial_1,
            evaluacion.calificacion_parcial_2,
            evaluacion.nota_final,
            evaluacion.porcentaje_asistencia,
            evaluacion.estado_de_aprobacion,
        ])

    documento_excel = HttpResponse(content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") #Indicar documento .xlsx
    documento_excel["Content-Disposition"] = f'attachment; filename="informe_general_{periodo.periodo}.xlsx"' #Descarga de documento
    documento_en_memoria.save(documento_excel)
    return documento_excel


def _exportar_informe_pdf(periodo, evaluaciones): #Crear documento de texto plano (.txt)
    lineas = [
        f"INFORME GENERAL DE NIVELACIÓN",
        f"Periodo de nivelación: {periodo.periodo}",
        f"Universidad: {periodo.universidad.nombre}",
        f"Fecha de generación: {date.today()}",
        f"{'─' * 60}",
        f"{'Número de identificación':<15} {'Nombres':<25} {'Unidad curricular':<30} {'Nota final':>6} {'Estado':<12}",
        f"{'─' * 60}",
    ]
    for evaluacion in evaluaciones:
        estudiante_actual = evaluacion.estudiante
        lineas.append(
            f"{estudiante_actual.usuario_de_sistema.identificacion:<15} "
            f"{estudiante_actual.usuario_de_sistema.nombres:<25} "
            f"{evaluacion.unidad_curricular.nombre:<30} "
            f"{evaluacion.nota_final:>6.2f} "
            f"{evaluacion.estado_de_aprobacion:<12}"
        )

    documento = "\n".join(lineas)
    documento_txt = HttpResponse(documento, content_type = "text/plain; charset=utf-8")
    documento_txt["Content-Disposition"] = f'attachment; filename="informe_general_{periodo.periodo}.txt"'
    return documento_txt